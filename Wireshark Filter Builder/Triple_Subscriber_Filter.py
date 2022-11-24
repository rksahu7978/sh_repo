# Creating static filter for 3 subscribers.
# The resulting static filter can be used for both 4G LTE core and VoLTE traces.

def sip_ctn(ctn1, ctn2, ctn3):
    return f'sip contains 44{ctn1} or sip contains 44{ctn2} or sip contains 44{ctn3}'


def sip_imsi(imsi1, imsi2, imsi3):
    return f'sip contains {imsi1} or sip contains {imsi2} or sip contains {imsi3}'


def diameter_ctn(ctn1, ctn2, ctn3):
    return f'diameter contains 44{ctn1} or diameter contains 44{ctn2} or diameter contains 44{ctn3}'


def diameter_imsi(imsi1, imsi2, imsi3):
    return f'diameter contains {imsi1} or diameter contains {imsi2} or diameter contains {imsi3}'


def e164(ctn1, ctn2, ctn3):
    return f'e164.msisdn == "44{ctn1}" or e164.msisdn == "44{ctn2}" or e164.msisdn == "44{ctn3}"'


def e212(imsi1, imsi2, imsi3):
    return f'e212.imsi == "{imsi1}" or e212.imsi == "{imsi2}" or e212.imsi == "{imsi3}"'


def isup_calling(ctn1, ctn2, ctn3):
    return f'isup.calling == "44{ctn1}" or isup.calling == "44{ctn2}" or isup.calling == "44{ctn3}"'


def isup_called(ctn1, ctn2, ctn3):
    return f'isup.called == "0{ctn1}" or isup.called == "0{ctn2}" or isup.called == "0{ctn3}"'


def dns(ctn1, ctn2, ctn3):
    temp1 = ""
    for digit in (str(ctn1)[::-1] + "44"):
        temp1 = temp1 + f'{digit}.'

    temp2 = ""
    for digit in (str(ctn2)[::-1] + "44"):
        temp2 = temp2 + f'{digit}.'

    temp3 = ""
    for digit in (str(ctn3)[::-1] + "44"):
        temp3 = temp3 + f'{digit}.'

    return f'dns.qry.name == "{temp1}e164.arpa" or dns.qry.name == "{temp2}e164.arpa" or dns.qry.name == "{temp3}e164.arpa"'


def get_msisdn(row):
    from openpyxl import load_workbook
    wb = load_workbook('.\\Sub_Details.xlsx')
    ws = wb.active
    msisdn = ws[f'A{row}'].value
    return msisdn


def get_imsi(row):
    from openpyxl import load_workbook
    wb = load_workbook('.\\Sub_Details.xlsx')
    ws = wb.active
    imsi = ws[f'B{row}'].value
    return imsi


def triple_filter(ctn1, ctn2, ctn3, imsi1, imsi2, imsi3):
    return f'{sip_ctn(ctn1, ctn2, ctn3)} or {sip_imsi(imsi1, imsi2, imsi3)} or {diameter_ctn(ctn1, ctn2, ctn3)} or {diameter_imsi(imsi1, imsi2, imsi3)} or {e164(ctn1, ctn2, ctn3)} or {e212(imsi1, imsi2, imsi3)} or {isup_calling(ctn1, ctn2, ctn3)} or {isup_called(ctn1, ctn2, ctn3)} or {dns(ctn1, ctn2, ctn3)} '


CTN1 = get_msisdn(input("Enter the row for 1st MSISDN :- "))
IMSI1 = get_imsi(input("Enter the row for 1st imsi :- "))
CTN2 = get_msisdn(input("Enter the row for 2nd MSISDN :- "))
IMSI2 = get_imsi(input("Enter the row for 2nd imsi :- "))
CTN3 = get_msisdn(input("Enter the row for 3rd MSISDN :- "))
IMSI3 = get_imsi(input("Enter the row for 3rd imsi :- "))

print(triple_filter(CTN1, CTN2, CTN3, IMSI1, IMSI2, IMSI3))
