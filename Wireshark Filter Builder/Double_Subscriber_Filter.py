# Creating static filter for 2 subscribers.
# The resulting static filter can be used for both 4G LTE core and VoLTE traces.

def sip(mo_ctn, mt_ctn):
    return f'sip contains 44{mo_ctn} or sip contains 44{mt_ctn}'


def diameter(mo_ctn, mt_ctn):
    return f'diameter contains 44{mo_ctn} or diameter contains 44{mt_ctn}'


def isup(mo_ctn, mt_ctn):
    return f'isup.calling == "44{mo_ctn}" or isup.called == "0{mt_ctn}" or isup.calling == "44{mt_ctn}" or isup.called == "0{mo_ctn}"'


def dns(mo_ctn, mt_ctn):
    temp1 = ""
    for digit in (str(mt_ctn)[::-1] + "44"):
        temp1 = temp1 + f'{digit}.'

    temp2 = ""
    for digit in (str(mo_ctn)[::-1] + "44"):
        temp2 = temp2 + f'{digit}.'

    return f'dns.qry.name == "{temp1}e164.arpa" or dns.qry.name == "{temp2}e164.arpa"'


def get_msisdn(row):
    from openpyxl import load_workbook
    wb = load_workbook('.\\Sub_Details.xlsx')
    ws = wb.active
    msisdn = ws[f'A{row}'].value
    return msisdn


def get_imsi(row):
    from openpyxl import load_workbook
    wb = load_workbook('.\\Sub_Details.xlsx')
    ws = wb.active
    imsi = ws[f'B{row}'].value
    return imsi


def sip_diameter_imsi(mo_imsi, mt_imsi):
    return f'sip contains {mo_imsi} or diameter contains {mo_imsi} or sip contains {mt_imsi} or diameter contains {mt_imsi}'


def double_filter(mo_ctn, mt_ctn, mo_imsi, mt_imsi):
    return f'{sip(mo_ctn, mt_ctn)} or {diameter(mo_ctn, mt_ctn)} or {sip_diameter_imsi(mo_imsi, mt_imsi)} or {isup(mo_ctn, mt_ctn)} or {dns(mo_ctn, mt_ctn)}'


MO_CTN = get_msisdn(input("Enter the row for MO MSISDN :- "))
MO_IMSI = get_imsi(input("Enter the row for MO IMSI :- "))
MT_CTN = get_msisdn(input("Enter the row for MT MSISDN :- "))
MT_IMSI = get_imsi(input("Enter the row for MT IMSI :- "))

print(double_filter(MO_CTN, MT_CTN, MO_IMSI, MT_IMSI))
