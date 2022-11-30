def get_msisdn(row):  # Getting MSISDN
    from openpyxl import load_workbook
    wb = load_workbook('.\\Sub_Details.xlsx')
    ws = wb.active
    return ws[f'A{row}'].value


def get_imsi(row):  # Getting IMSI
    from openpyxl import load_workbook
    wb = load_workbook('.\\Sub_Details.xlsx')
    ws = wb.active
    return ws[f'B{row}'].value


def get_list(num_of_cycle):  # Getting list of subscribers
    ctn_list = []
    imsi_list = []
    count = 1
    for num in range(1, num_of_cycle + 1, 1):
        if num == 1:
            row_num = input(f'Enter the {count}st row number :- ')
            ctn = str(get_msisdn(row_num))
            imsi = get_imsi(row_num)
        elif num == 2:
            row_num = input(f'Enter the {count}nd row number :- ')
            ctn = str(get_msisdn(row_num))
            imsi = get_imsi(row_num)
        elif num == 3:
            row_num = input(f'Enter the {count}rd row number :- ')
            ctn = str(get_msisdn(row_num))
            imsi = get_imsi(row_num)
        else:
            row_num = input(f'Enter the {count}th row number :- ')
            ctn = str(get_msisdn(row_num))
            imsi = get_imsi(row_num)
        count += 1
        ctn_list.append(ctn)
        imsi_list.append(imsi)
    return ctn_list, imsi_list


# Building filters with MSISDN

def sip_ctn(ctn_list):  # SIP Protocol
    sip_ctn_final = ""
    for ctn in ctn_list:
        sip_ctn_final = f'sip contains 44{ctn} or ' + sip_ctn_final
    return sip_ctn_final[0:len(sip_ctn_final) - 4:]


def diameter_ctn(ctn_list):  # DIAMETER Protocol
    diameter_ctn_final = ""
    for ctn in ctn_list:
        diameter_ctn_final = f'diameter contains 44{ctn} or ' + diameter_ctn_final
    return diameter_ctn_final[0:len(diameter_ctn_final) - 4:]


def isup_calling(ctn_list):  # CAMEL Protocol
    isup_calling_final = ""
    for ctn in ctn_list:
        isup_calling_final = f'isup.calling == "44{ctn}" or ' + isup_calling_final
    return isup_calling_final[0:len(isup_calling_final) - 4:]


def isup_called(ctn_list):  # CAMEL Protocol
    isup_called_final = ""
    for ctn in ctn_list:
        isup_called_final = f'isup.called == "0{ctn}" or ' + isup_called_final
    return isup_called_final[0:len(isup_called_final) - 4:]


def e164(ctn_list):  # E.164 Format
    e164_final = ""
    for ctn in ctn_list:
        e164_final = f'e164.msisdn == "44{ctn}" or ' + e164_final
    return e164_final[0:len(e164_final) - 4:]


def enum(ctn_list):  # DNS Protocol
    enum_final = ""
    for ctn in ctn_list:
        temp1 = ""
        for digit in (ctn[::-1] + "44"):
            temp1 = temp1 + f'{digit}.'
        enum_final = f'dns.qry.name == "{temp1}e164.arpa" or ' + enum_final
    return enum_final[0:len(enum_final) - 4:]


def http_ctn(ctn_list):  # HTTP protocol
    http_ctn_final = ""
    for ctn in ctn_list:
        http_ctn_final = f'http.request.uri contains 44{ctn} or http.response_for.uri contains 44{ctn} or xml.cdata contains 44{ctn} or ' + http_ctn_final
    return http_ctn_final[0:len(http_ctn_final) - 4:]

# Building filters with IMSI


def sip_imsi(imsi_list):  # SIP Protocol
    sip_imsi_final = ""
    for imsi in imsi_list:
        sip_imsi_final = f'sip contains {imsi} or ' + sip_imsi_final
    return sip_imsi_final[0:len(sip_imsi_final) - 4:]


def diameter_imsi(imsi_list):  # DIAMETER Protocol
    diameter_imsi_final = ""
    for imsi in imsi_list:
        diameter_imsi_final = f'diameter contains {imsi} or ' + diameter_imsi_final
    return diameter_imsi_final[0:len(diameter_imsi_final) - 4:]


def e212(imsi_list):  # E.212 Format
    e212_final = ""
    for imsi in imsi_list:
        e212_final = f'e212.imsi == "{imsi}" or ' + e212_final
    return e212_final[0:len(e212_final) - 4:]


def http_imsi(imsi_list):  # HTTP Protocol
    http_imsi_final = ""
    for imsi in imsi_list:
        http_imsi_final = f'http.authorization contains {imsi} or http.response_for.uri contains {imsi} or ' + http_imsi_final
    return http_imsi_final[0:len(http_imsi_final) - 4:]


class Master:  # Master class
    def __init__(self, net_sub):
        self.net_sub = net_sub


master_obj = Master(int(input("Enter the total number of subscribers :- ")))

net_list = list(get_list(master_obj.net_sub))

print(f'{sip_ctn(net_list[0])} or {diameter_ctn(net_list[0])} or {isup_calling(net_list[0])} or {isup_called(net_list[0])} or {e164(net_list[0])} or {enum(net_list[0])} or {http_ctn(net_list[0])} or {sip_imsi(net_list[1])} or {diameter_imsi(net_list[1])} or {e212(net_list[1])} or {http_imsi(net_list[1])}')
