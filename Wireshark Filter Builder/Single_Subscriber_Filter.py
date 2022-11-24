# Creating static filter for a single subscriber.
# The resulting static filter can be used for both 4G LTE core and VoLTE traces.

def sip(msisdn, imsi):
    return f'sip contains 44{msisdn} or sip contains {imsi}'


def diameter(msisdn, imsi):
    return f'diameter contains 44{msisdn} or diameter contains {imsi}'


def e164_e212(msisdn, imsi):
    return f'e164.msisdn == "44{msisdn}" or e212.imsi == "{imsi}"'


def single_filter(msisdn, imsi):
    return f'{sip(msisdn, imsi)} or {diameter(msisdn, imsi)} or {e164_e212(msisdn, imsi)}'


def get_msisdn(row):
    from openpyxl import load_workbook
    wb = load_workbook('.\\Sub_Details.xlsx')
    ws = wb.active
    value = ws[f'A{row}'].value
    return value


def get_imsi(row):
    from openpyxl import load_workbook
    wb = load_workbook('.\\Sub_Details.xlsx')
    ws = wb.active
    value = ws[f'B{row}'].value
    return value


MSISDN = get_msisdn(input("Enter the row for MSISDN :- "))
IMSI = get_imsi(input("Enter the row for IMSI :- "))

print(single_filter(MSISDN, IMSI))
